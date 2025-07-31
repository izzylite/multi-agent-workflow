"""
Enhanced Export Manager for CLI scraping results.

This module provides a comprehensive export management system with capabilities
for exporting results in multiple formats with advanced configuration options,
filtering, sorting, and progress tracking.
"""

import json
import gzip
import csv
import zipfile
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any, Optional, Union, Tuple, Callable
from dataclasses import dataclass, asdict
from enum import Enum

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


class ExportFormat(Enum):
    """Supported export formats."""
    JSON = "json"
    CSV = "csv"
    EXCEL = "excel"
    MARKDOWN = "markdown"


class CompressionFormat(Enum):
    """Supported compression formats."""
    NONE = "none"
    GZIP = "gzip"
    ZIP = "zip"


@dataclass
class ExportConfig:
    """Configuration for export operations."""
    format: ExportFormat
    compression: CompressionFormat = CompressionFormat.NONE
    pretty_print: bool = True
    include_metadata: bool = True
    max_file_size: Optional[int] = None
    chunk_size: int = 1000
    progress_callback: Optional[Callable[[int, int, str], None]] = None


@dataclass
class CSVConfig:
    """Configuration for CSV export."""
    delimiter: str = ","
    quotechar: str = '"'
    quoting: int = csv.QUOTE_MINIMAL
    encoding: str = "utf-8"
    include_header: bool = True


@dataclass
class ExcelConfig:
    """Configuration for Excel export."""
    sheet_name: str = "Products"
    auto_filter: bool = True
    freeze_panes: bool = True
    header_style: bool = True
    column_widths: Optional[Dict[str, int]] = None
    number_format: str = "#,##0"


@dataclass
class MarkdownConfig:
    """Configuration for Markdown export."""
    include_table: bool = True
    include_summary: bool = True
    max_rows: Optional[int] = None
    sort_by: Optional[str] = None
    sort_ascending: bool = True


@dataclass
class ExportTemplate:
    """Export template for common use cases."""
    name: str
    description: str
    config: ExportConfig
    fields: Optional[List[str]] = None
    filters: Optional[Dict[str, Any]] = None
    sort_by: Optional[str] = None
    sort_ascending: bool = True
    csv_config: Optional[CSVConfig] = None
    excel_config: Optional[ExcelConfig] = None
    markdown_config: Optional[MarkdownConfig] = None


class ExportManager:
    """
    Enhanced export manager with multi-format support and advanced features.
    
    Provides export capabilities for JSON, CSV, Excel, and Markdown formats
    with configuration options, filtering, sorting, and progress tracking.
    """

    def __init__(self, output_dir: str = "exports"):
        """
        Initialize the ExportManager.
        
        Args:
            output_dir: Directory for exported files
        """
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.templates: Dict[str, ExportTemplate] = {}
        self._load_default_templates()

    def _load_default_templates(self):
        """Load default export templates."""
        # Summary template
        summary_template = ExportTemplate(
            name="summary",
            description="Export summary statistics only",
            config=ExportConfig(format=ExportFormat.JSON, pretty_print=True),
            fields=["vendor", "category", "created_at"],
            sort_by="created_at",
            sort_ascending=False
        )
        self.templates["summary"] = summary_template
        
        # Detailed template
        detailed_template = ExportTemplate(
            name="detailed",
            description="Export all fields with detailed formatting",
            config=ExportConfig(format=ExportFormat.EXCEL, pretty_print=True),
            excel_config=ExcelConfig(
                sheet_name="Detailed Data",
                auto_filter=True,
                freeze_panes=True,
                header_style=True
            )
        )
        self.templates["detailed"] = detailed_template
        
        # CSV template
        csv_template = ExportTemplate(
            name="csv_standard",
            description="Standard CSV export with UTF-8 encoding",
            config=ExportConfig(format=ExportFormat.CSV),
            csv_config=CSVConfig(
                delimiter=",",
                encoding="utf-8",
                include_header=True
            )
        )
        self.templates["csv_standard"] = csv_template
        
        # Compressed template
        compressed_template = ExportTemplate(
            name="compressed",
            description="Compressed export for large datasets",
            config=ExportConfig(
                format=ExportFormat.JSON,
                compression=CompressionFormat.GZIP,
                pretty_print=False
            )
        )
        self.templates["compressed"] = compressed_template

    def export_data(
        self,
        data: List[Dict[str, Any]],
        output_path: str,
        config: ExportConfig,
        fields: Optional[List[str]] = None,
        filters: Optional[Dict[str, Any]] = None,
        sort_by: Optional[str] = None,
        sort_ascending: bool = True,
        csv_config: Optional[CSVConfig] = None,
        excel_config: Optional[ExcelConfig] = None,
        markdown_config: Optional[MarkdownConfig] = None
    ) -> str:
        """
        Export data to the specified format with advanced options.
        
        Args:
            data: Data to export
            output_path: Output file path
            config: Export configuration
            fields: Specific fields to export
            filters: Filter conditions
            sort_by: Field to sort by
            sort_ascending: Sort order
            csv_config: CSV-specific configuration
            excel_config: Excel-specific configuration
            markdown_config: Markdown-specific configuration
            
        Returns:
            Path to the exported file
        """
        if not data:
            raise ValueError("No data to export")
        
        # Convert to DataFrame
        df = pd.DataFrame(data)
        
        # Apply filters
        if filters:
            df = self._apply_filters(df, filters)
        
        # Apply field selection
        if fields:
            df = self._select_fields(df, fields)
        
        # Apply sorting
        if sort_by and sort_by in df.columns:
            df = df.sort_values(by=sort_by, ascending=sort_ascending)
        
        # Export based on format
        if config.format == ExportFormat.JSON:
            return self._export_json(df, output_path, config)
        elif config.format == ExportFormat.CSV:
            return self._export_csv(df, output_path, config, csv_config)
        elif config.format == ExportFormat.EXCEL:
            return self._export_excel(df, output_path, config, excel_config)
        elif config.format == ExportFormat.MARKDOWN:
            return self._export_markdown(df, output_path, config, markdown_config)
        else:
            raise ValueError(f"Unsupported export format: {config.format}")

    def export_result_file(
        self,
        file_path: str,
        output_path: str,
        config: ExportConfig,
        fields: Optional[List[str]] = None,
        filters: Optional[Dict[str, Any]] = None,
        sort_by: Optional[str] = None,
        sort_ascending: bool = True
    ) -> str:
        """
        Export a result file to the specified format.
        
        Args:
            file_path: Path to the result file
            output_path: Output file path
            config: Export configuration
            fields: Specific fields to export
            filters: Filter conditions
            sort_by: Field to sort by
            sort_ascending: Sort order
            
        Returns:
            Path to the exported file
        """
        # Load the result data
        data = self._load_result_file(file_path)
        
        return self.export_data(
            data, output_path, config, fields, filters, sort_by, sort_ascending
        )

    def batch_export(
        self,
        file_paths: List[str],
        output_dir: str,
        config: ExportConfig,
        fields: Optional[List[str]] = None,
        filters: Optional[Dict[str, Any]] = None,
        sort_by: Optional[str] = None,
        sort_ascending: bool = True
    ) -> List[str]:
        """
        Export multiple result files in batch.
        
        Args:
            file_paths: List of file paths to export
            output_dir: Output directory
            config: Export configuration
            fields: Specific fields to export
            filters: Filter conditions
            sort_by: Field to sort by
            sort_ascending: Sort order
            
        Returns:
            List of exported file paths
        """
        exported_files = []
        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)
        
        total_files = len(file_paths)
        for i, file_path in enumerate(file_paths):
            if config.progress_callback:
                config.progress_callback(i + 1, total_files, f"Exporting {Path(file_path).name}")
            
            # Generate output filename
            input_path = Path(file_path)
            output_filename = f"{input_path.stem}.{config.format.value}"
            output_path = output_dir / output_filename
            
            try:
                exported_path = self.export_result_file(
                    file_path, str(output_path), config, fields, filters, sort_by, sort_ascending
                )
                exported_files.append(exported_path)
            except Exception as e:
                print(f"Failed to export {file_path}: {e}")
                continue
        
        return exported_files

    def incremental_export(
        self,
        data: List[Dict[str, Any]],
        output_path: str,
        config: ExportConfig,
        chunk_size: int = 1000,
        resume_from: Optional[int] = 0,
        fields: Optional[List[str]] = None,
        filters: Optional[Dict[str, Any]] = None,
        sort_by: Optional[str] = None,
        sort_ascending: bool = True
    ) -> str:
        """
        Export data incrementally in chunks for large datasets.
        
        Args:
            data: Data to export
            output_path: Output file path
            config: Export configuration
            chunk_size: Number of records per chunk
            resume_from: Index to resume from (for interrupted exports)
            fields: Specific fields to export
            filters: Filter conditions
            sort_by: Field to sort by
            sort_ascending: Sort order
            
        Returns:
            Path to the exported file
        """
        if not data:
            raise ValueError("No data to export")
        
        # Convert to DataFrame
        df = pd.DataFrame(data)
        
        # Apply filters
        if filters:
            df = self._apply_filters(df, filters)
        
        # Apply field selection
        if fields:
            df = self._select_fields(df, fields)
        
        # Apply sorting
        if sort_by and sort_by in df.columns:
            df = df.sort_values(by=sort_by, ascending=sort_ascending)
        
        total_records = len(df)
        output_path = Path(output_path)
        
        # Handle different formats for incremental export
        if config.format == ExportFormat.JSON:
            return self._incremental_export_json(df, output_path, config, chunk_size, resume_from)
        elif config.format == ExportFormat.CSV:
            return self._incremental_export_csv(df, output_path, config, chunk_size, resume_from)
        elif config.format == ExportFormat.EXCEL:
            return self._incremental_export_excel(df, output_path, config, chunk_size, resume_from)
        elif config.format == ExportFormat.MARKDOWN:
            return self._incremental_export_markdown(df, output_path, config, chunk_size, resume_from)
        else:
            raise ValueError(f"Unsupported export format for incremental export: {config.format}")

    def _incremental_export_json(
        self,
        df: pd.DataFrame,
        output_path: Path,
        config: ExportConfig,
        chunk_size: int,
        resume_from: int
    ) -> str:
        """Incremental JSON export."""
        total_records = len(df)
        
        # Initialize export data structure
        export_data = {
            "metadata": {
                "exported_at": datetime.now().isoformat(),
                "total_records": total_records,
                "fields": df.columns.tolist(),
                "format": "json",
                "incremental": True,
                "chunk_size": chunk_size
            },
            "data": []
        }
        
        # Process chunks
        for start_idx in range(resume_from, total_records, chunk_size):
            end_idx = min(start_idx + chunk_size, total_records)
            chunk_df = df.iloc[start_idx:end_idx]
            
            # Add chunk data
            export_data["data"].extend(chunk_df.to_dict('records'))
            
            # Update progress
            if config.progress_callback:
                config.progress_callback(end_idx, total_records, f"Processing chunk {start_idx//chunk_size + 1}")
            
            # Write intermediate file (for resume capability)
            with open(output_path, 'w', encoding='utf-8') as f:
                if config.pretty_print:
                    json.dump(export_data, f, indent=2, ensure_ascii=False)
                else:
                    json.dump(export_data, f, ensure_ascii=False)
        
        # Apply compression if requested
        if config.compression == CompressionFormat.GZIP:
            return self._compress_file(output_path, 'gzip')
        elif config.compression == CompressionFormat.ZIP:
            return self._compress_file(output_path, 'zip')
        
        return str(output_path)

    def _incremental_export_csv(
        self,
        df: pd.DataFrame,
        output_path: Path,
        config: ExportConfig,
        chunk_size: int,
        resume_from: int
    ) -> str:
        """Incremental CSV export."""
        total_records = len(df)
        
        # Write header first
        df.head(0).to_csv(output_path, index=False, mode='w')
        
        # Process chunks
        for start_idx in range(resume_from, total_records, chunk_size):
            end_idx = min(start_idx + chunk_size, total_records)
            chunk_df = df.iloc[start_idx:end_idx]
            
            # Append chunk to CSV
            chunk_df.to_csv(output_path, index=False, mode='a', header=False)
            
            # Update progress
            if config.progress_callback:
                config.progress_callback(end_idx, total_records, f"Processing chunk {start_idx//chunk_size + 1}")
        
        # Apply compression if requested
        if config.compression == CompressionFormat.GZIP:
            return self._compress_file(output_path, 'gzip')
        elif config.compression == CompressionFormat.ZIP:
            return self._compress_file(output_path, 'zip')
        
        return str(output_path)

    def _incremental_export_excel(
        self,
        df: pd.DataFrame,
        output_path: Path,
        config: ExportConfig,
        chunk_size: int,
        resume_from: int
    ) -> str:
        """Incremental Excel export."""
        total_records = len(df)
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Products"
        
        # Write header
        for col_idx, col_name in enumerate(df.columns, 1):
            ws.cell(row=1, column=col_idx, value=col_name)
        
        # Process chunks
        row_idx = 2  # Start after header
        for start_idx in range(resume_from, total_records, chunk_size):
            end_idx = min(start_idx + chunk_size, total_records)
            chunk_df = df.iloc[start_idx:end_idx]
            
            # Write chunk data
            for _, row in chunk_df.iterrows():
                for col_idx, value in enumerate(row, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
                row_idx += 1
            
            # Update progress
            if config.progress_callback:
                config.progress_callback(end_idx, total_records, f"Processing chunk {start_idx//chunk_size + 1}")
        
        # Apply formatting
        if hasattr(self, '_format_excel_header'):
            self._format_excel_header(ws, len(df.columns))
        
        # Save workbook
        wb.save(output_path)
        
        return str(output_path)

    def _incremental_export_markdown(
        self,
        df: pd.DataFrame,
        output_path: Path,
        config: ExportConfig,
        chunk_size: int,
        resume_from: int
    ) -> str:
        """Incremental Markdown export."""
        total_records = len(df)
        
        # Write header
        lines = [
            "# Data Export (Incremental)",
            "",
            f"**Exported:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            f"**Total Records:** {total_records}",
            f"**Fields:** {', '.join(df.columns.tolist())}",
            f"**Chunk Size:** {chunk_size}",
            ""
        ]
        
        # Write header to file
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write('\n'.join(lines))
        
        # Process chunks
        for start_idx in range(resume_from, total_records, chunk_size):
            end_idx = min(start_idx + chunk_size, total_records)
            chunk_df = df.iloc[start_idx:end_idx]
            
            # Convert chunk to markdown table
            chunk_lines = self._dataframe_to_markdown_table(chunk_df)
            
            # Append chunk to file
            with open(output_path, 'a', encoding='utf-8') as f:
                f.write('\n'.join(chunk_lines))
                f.write('\n')
            
            # Update progress
            if config.progress_callback:
                config.progress_callback(end_idx, total_records, f"Processing chunk {start_idx//chunk_size + 1}")
        
        return str(output_path)

    def validate_export(
        self,
        exported_path: str,
        original_data: List[Dict[str, Any]],
        fields: Optional[List[str]] = None
    ) -> Dict[str, Any]:
        """
        Validate exported data for integrity.
        
        Args:
            exported_path: Path to exported file
            original_data: Original data for comparison
            fields: Fields that were exported
            
        Returns:
            Validation results
        """
        validation_results = {
            "file_exists": False,
            "file_size": 0,
            "row_count": 0,
            "field_count": 0,
            "data_integrity": False,
            "errors": []
        }
        
        try:
            # Check file exists
            file_path = Path(exported_path)
            validation_results["file_exists"] = file_path.exists()
            
            if not validation_results["file_exists"]:
                validation_results["errors"].append("Exported file does not exist")
                return validation_results
            
            # Get file size
            validation_results["file_size"] = file_path.stat().st_size
            
            # Validate based on format
            if exported_path.endswith('.json'):
                validation_results.update(self._validate_json(file_path, original_data, fields))
            elif exported_path.endswith('.csv'):
                validation_results.update(self._validate_csv(file_path, original_data, fields))
            elif exported_path.endswith('.xlsx'):
                validation_results.update(self._validate_excel(file_path, original_data, fields))
            elif exported_path.endswith('.md'):
                validation_results.update(self._validate_markdown(file_path, original_data, fields))
            
        except Exception as e:
            validation_results["errors"].append(f"Validation error: {e}")
        
        return validation_results

    def _export_json(
        self,
        df: pd.DataFrame,
        output_path: str,
        config: ExportConfig
    ) -> str:
        """Export data to JSON format."""
        output_path = Path(output_path)
        
        # Prepare export data
        export_data = {
            "metadata": {
                "exported_at": datetime.now().isoformat(),
                "total_records": len(df),
                "fields": df.columns.tolist(),
                "format": "json"
            },
            "data": df.to_dict('records')
        }
        
        # Write JSON
        with open(output_path, 'w', encoding='utf-8') as f:
            if config.pretty_print:
                json.dump(export_data, f, indent=2, ensure_ascii=False)
            else:
                json.dump(export_data, f, ensure_ascii=False)
        
        # Update progress
        if config.progress_callback:
            config.progress_callback(len(df), len(df), "Export completed")
        
        # Apply compression if requested
        if config.compression == CompressionFormat.GZIP:
            return self._compress_file(output_path, 'gzip')
        elif config.compression == CompressionFormat.ZIP:
            return self._compress_file(output_path, 'zip')
        
        return str(output_path)

    def _export_csv(
        self,
        df: pd.DataFrame,
        output_path: str,
        config: ExportConfig,
        csv_config: Optional[CSVConfig] = None
    ) -> str:
        """Export data to CSV format."""
        if csv_config is None:
            csv_config = CSVConfig()
        
        output_path = Path(output_path)
        
        # Write CSV
        df.to_csv(
            output_path,
            index=False,
            sep=csv_config.delimiter,
            quotechar=csv_config.quotechar,
            quoting=csv_config.quoting,
            encoding=csv_config.encoding,
            header=csv_config.include_header
        )
        
        # Update progress
        if config.progress_callback:
            config.progress_callback(len(df), len(df), "Export completed")
        
        # Apply compression if requested
        if config.compression == CompressionFormat.GZIP:
            return self._compress_file(output_path, 'gzip')
        elif config.compression == CompressionFormat.ZIP:
            return self._compress_file(output_path, 'zip')
        
        return str(output_path)

    def _export_excel(
        self,
        df: pd.DataFrame,
        output_path: str,
        config: ExportConfig,
        excel_config: Optional[ExcelConfig] = None
    ) -> str:
        """Export data to Excel format with formatting."""
        if excel_config is None:
            excel_config = ExcelConfig()
        
        output_path = Path(output_path)
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = excel_config.sheet_name
        
        # Write data
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # Apply formatting
        if excel_config.header_style:
            self._format_excel_header(ws, len(df.columns))
        
        # Set column widths
        if excel_config.column_widths:
            for col, width in excel_config.column_widths.items():
                if col in df.columns:
                    col_idx = df.columns.get_loc(col) + 1
                    ws.column_dimensions[chr(64 + col_idx)].width = width
        
        # Apply auto filter
        if excel_config.auto_filter:
            ws.auto_filter.ref = ws.dimensions
        
        # Freeze panes
        if excel_config.freeze_panes:
            ws.freeze_panes = 'A2'
        
        # Save workbook
        wb.save(output_path)
        
        return str(output_path)

    def _export_markdown(
        self,
        df: pd.DataFrame,
        output_path: str,
        config: ExportConfig,
        markdown_config: Optional[MarkdownConfig] = None
    ) -> str:
        """Export data to Markdown format."""
        if markdown_config is None:
            markdown_config = MarkdownConfig()
        
        output_path = Path(output_path)
        
        lines = []
        
        # Add header
        lines.append(f"# Data Export")
        lines.append(f"")
        lines.append(f"**Exported:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        lines.append(f"**Total Records:** {len(df)}")
        lines.append(f"**Fields:** {', '.join(df.columns.tolist())}")
        lines.append(f"")
        
        # Add summary if requested
        if markdown_config.include_summary:
            lines.extend(self._generate_markdown_summary(df))
        
        # Add table if requested
        if markdown_config.include_table:
            # Limit rows if specified
            display_df = df
            if markdown_config.max_rows and len(df) > markdown_config.max_rows:
                display_df = df.head(markdown_config.max_rows)
                lines.append(f"*Showing first {markdown_config.max_rows} records*")
                lines.append(f"")
            
            # Convert to markdown table
            table_lines = self._dataframe_to_markdown_table(display_df)
            lines.extend(table_lines)
        
        # Write markdown
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write('\n'.join(lines))
        
        return str(output_path)

    def _apply_filters(self, df: pd.DataFrame, filters: Dict[str, Any]) -> pd.DataFrame:
        """Apply filters to DataFrame."""
        for field, condition in filters.items():
            if field in df.columns:
                if isinstance(condition, dict):
                    # Complex filter (e.g., {"operator": "contains", "value": "text"})
                    operator = condition.get("operator", "eq")
                    value = condition.get("value")
                    
                    if operator == "contains":
                        df = df[df[field].astype(str).str.contains(str(value), na=False)]
                    elif operator == "startswith":
                        df = df[df[field].astype(str).str.startswith(str(value), na=False)]
                    elif operator == "endswith":
                        df = df[df[field].astype(str).str.endswith(str(value), na=False)]
                    elif operator == "in":
                        df = df[df[field].isin(value)]
                    elif operator == "between":
                        df = df[df[field].between(condition.get("min"), condition.get("max"))]
                else:
                    # Simple equality filter
                    df = df[df[field] == condition]
        
        return df

    def _select_fields(self, df: pd.DataFrame, fields: List[str]) -> pd.DataFrame:
        """Select specific fields from DataFrame."""
        available_fields = df.columns.tolist()
        valid_fields = [f for f in fields if f in available_fields]
        
        if not valid_fields:
            raise ValueError(f"No valid fields found. Available: {available_fields}")
        
        return df[valid_fields]

    def _load_result_file(self, file_path: Union[str, Path]) -> List[Dict[str, Any]]:
        """Load result file data."""
        file_path = Path(file_path)
        
        if not file_path.exists():
            raise FileNotFoundError(f"Result file not found: {file_path}")
        
        # Handle compressed files
        if file_path.suffix == '.gz':
            with gzip.open(file_path, 'rt', encoding='utf-8') as f:
                return json.load(f)
        else:
            with open(file_path, 'r', encoding='utf-8') as f:
                return json.load(f)

    def _compress_file(self, file_path: Path, compression_type: str) -> str:
        """Compress a file and return the compressed path."""
        if compression_type == 'gzip':
            compressed_path = file_path.with_suffix(file_path.suffix + '.gz')
            with open(file_path, 'rb') as f_in:
                with gzip.open(compressed_path, 'wb') as f_out:
                    f_out.writelines(f_in)
            file_path.unlink()  # Remove original file
            return str(compressed_path)
        elif compression_type == 'zip':
            compressed_path = file_path.with_suffix(file_path.suffix + '.zip')
            with zipfile.ZipFile(compressed_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                zipf.write(file_path, file_path.name)
            file_path.unlink()  # Remove original file
            return str(compressed_path)
        
        return str(file_path)

    def _format_excel_header(self, ws, num_columns: int):
        """Format Excel header row."""
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col in range(1, num_columns + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

    def _generate_markdown_summary(self, df: pd.DataFrame) -> List[str]:
        """Generate markdown summary section."""
        lines = []
        lines.append("## Summary")
        lines.append("")
        
        # Basic stats
        lines.append(f"- **Total Records:** {len(df)}")
        lines.append(f"- **Columns:** {len(df.columns)}")
        lines.append("")
        
        # Column information
        lines.append("### Column Information")
        lines.append("")
        for col in df.columns:
            non_null_count = df[col].count()
            null_count = len(df) - non_null_count
            lines.append(f"- **{col}:** {non_null_count} non-null values, {null_count} null values")
        
        lines.append("")
        return lines

    def _dataframe_to_markdown_table(self, df: pd.DataFrame) -> List[str]:
        """Convert DataFrame to markdown table."""
        lines = []
        
        # Header
        lines.append("## Data")
        lines.append("")
        
        # Table header
        header = "| " + " | ".join(str(col) for col in df.columns) + " |"
        lines.append(header)
        
        # Separator
        separator = "| " + " | ".join("---" for _ in df.columns) + " |"
        lines.append(separator)
        
        # Data rows
        for _, row in df.iterrows():
            row_str = "| " + " | ".join(str(val) for val in row) + " |"
            lines.append(row_str)
        
        lines.append("")
        return lines

    def _validate_json(self, file_path: Path, original_data: List[Dict[str, Any]], fields: Optional[List[str]]) -> Dict[str, Any]:
        """Validate JSON export."""
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            exported_data = data.get('data', [])
            
            return {
                "row_count": len(exported_data),
                "field_count": len(exported_data[0]) if exported_data else 0,
                "data_integrity": len(exported_data) == len(original_data)
            }
        except Exception as e:
            return {
                "row_count": 0,
                "field_count": 0,
                "data_integrity": False,
                "errors": [f"JSON validation error: {e}"]
            }

    def _validate_csv(self, file_path: Path, original_data: List[Dict[str, Any]], fields: Optional[List[str]]) -> Dict[str, Any]:
        """Validate CSV export."""
        try:
            df = pd.read_csv(file_path)
            
            return {
                "row_count": len(df),
                "field_count": len(df.columns),
                "data_integrity": len(df) == len(original_data)
            }
        except Exception as e:
            return {
                "row_count": 0,
                "field_count": 0,
                "data_integrity": False,
                "errors": [f"CSV validation error: {e}"]
            }

    def _validate_excel(self, file_path: Path, original_data: List[Dict[str, Any]], fields: Optional[List[str]]) -> Dict[str, Any]:
        """Validate Excel export."""
        try:
            df = pd.read_excel(file_path)
            
            return {
                "row_count": len(df),
                "field_count": len(df.columns),
                "data_integrity": len(df) == len(original_data)
            }
        except Exception as e:
            return {
                "row_count": 0,
                "field_count": 0,
                "data_integrity": False,
                "errors": [f"Excel validation error: {e}"]
            }

    def _validate_markdown(self, file_path: Path, original_data: List[Dict[str, Any]], fields: Optional[List[str]]) -> Dict[str, Any]:
        """Validate Markdown export."""
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                content = f.read()
            
            # Count table rows (simple validation)
            lines = content.split('\n')
            table_lines = [line for line in lines if line.startswith('|') and '---' not in line]
            
            return {
                "row_count": len(table_lines) - 1,  # Subtract header
                "field_count": 0,  # Not easily determined from markdown
                "data_integrity": len(table_lines) > 1  # Has at least header and one data row
            }
        except Exception as e:
            return {
                "row_count": 0,
                "field_count": 0,
                "data_integrity": False,
                "errors": [f"Markdown validation error: {e}"]
            }

    def add_template(self, template: ExportTemplate) -> None:
        """
        Add a new export template.
        
        Args:
            template: Export template to add
        """
        self.templates[template.name] = template

    def get_template(self, name: str) -> Optional[ExportTemplate]:
        """
        Get an export template by name.
        
        Args:
            name: Template name
            
        Returns:
            Export template or None if not found
        """
        return self.templates.get(name)

    def list_templates(self) -> List[ExportTemplate]:
        """
        List all available templates.
        
        Returns:
            List of export templates
        """
        return list(self.templates.values())

    def remove_template(self, name: str) -> bool:
        """
        Remove an export template.
        
        Args:
            name: Template name to remove
            
        Returns:
            True if template was removed, False if not found
        """
        if name in self.templates:
            del self.templates[name]
            return True
        return False

    def export_with_template(
        self,
        data: List[Dict[str, Any]],
        output_path: str,
        template_name: str,
        **kwargs
    ) -> str:
        """
        Export data using a predefined template.
        
        Args:
            data: Data to export
            output_path: Output file path
            template_name: Name of the template to use
            **kwargs: Additional parameters to override template settings
            
        Returns:
            Path to the exported file
        """
        template = self.get_template(template_name)
        if not template:
            raise ValueError(f"Template '{template_name}' not found")
        
        # Merge template settings with overrides
        config = template.config
        fields = kwargs.get('fields', template.fields)
        filters = kwargs.get('filters', template.filters)
        sort_by = kwargs.get('sort_by', template.sort_by)
        sort_ascending = kwargs.get('sort_ascending', template.sort_ascending)
        csv_config = kwargs.get('csv_config', template.csv_config)
        excel_config = kwargs.get('excel_config', template.excel_config)
        markdown_config = kwargs.get('markdown_config', template.markdown_config)
        
        return self.export_data(
            data, output_path, config, fields, filters, sort_by, sort_ascending,
            csv_config, excel_config, markdown_config
        )

    def export_result_file_with_template(
        self,
        file_path: str,
        output_path: str,
        template_name: str,
        **kwargs
    ) -> str:
        """
        Export a result file using a predefined template.
        
        Args:
            file_path: Path to the result file
            output_path: Output file path
            template_name: Name of the template to use
            **kwargs: Additional parameters to override template settings
            
        Returns:
            Path to the exported file
        """
        template = self.get_template(template_name)
        if not template:
            raise ValueError(f"Template '{template_name}' not found")
        
        # Load the result data
        data = self._load_result_file(file_path)
        
        # Merge template settings with overrides
        config = template.config
        fields = kwargs.get('fields', template.fields)
        filters = kwargs.get('filters', template.filters)
        sort_by = kwargs.get('sort_by', template.sort_by)
        sort_ascending = kwargs.get('sort_ascending', template.sort_ascending)
        
        return self.export_data(
            data, output_path, config, fields, filters, sort_by, sort_ascending
        )


def create_export_manager(output_dir: str = "exports") -> ExportManager:
    """
    Create an ExportManager instance.
    
    Args:
        output_dir: Directory for exported files
        
    Returns:
        ExportManager instance
    """
    return ExportManager(output_dir) 